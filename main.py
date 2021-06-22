from PyQt5 import QtWidgets, QtGui
from PyQt5 import QtCore
from PyQt5.QtGui import QMovie
from PyQt5.QtWidgets import QCheckBox, QFileDialog, QLabel, QMainWindow, QTableWidgetItem, QAbstractItemView, \
    QMessageBox, QWidget, QHeaderView
import openpyxl
from openpyxl import Workbook
import re
from PyQt5.QtCore import QMetaObject, QSize, QThread, pyqtSignal

from openpyxl.styles import Font, Alignment
from openpyxl.utils.exceptions import InvalidFileException

from form import Ui_MainWindow  # импорт нашего сгенерированного файла
from AboutForm import Ui_Dialog
from ChooseForm import Choose_Dialog
from ChooseFilter import Ui_Dialog_ChooseFilter
from otdel import Otdel
from shipment import Shipment

import sys
import os
import images_qr
import xlwings as xw
import analize

from dicts import lifetime, choose_position, choose_position_header, choose_position_header_evry_two


class MyThread(QThread):
    showMessageBox = QtCore.pyqtSignal(list)
    fihishThread = QtCore.pyqtSignal(str)
    send_mnozh = QtCore.pyqtSignal(set)

    def __init__(self, my_window, parent=None):
        super(MyThread, self).__init__()
        self.my_window = my_window
        self.pause = False
        self.mnozh = set()

    def run(self):
        self.my_window.ui.statusbar.showMessage(
            'Анализ и сопоставление данных исходной таблицы')
        self.my_window.ui.pushButton_2.setEnabled(False)
        self.my_window.ui.pushButton_3.setEnabled(False)
        wb = openpyxl.load_workbook(self.my_window.filename[0])
        sheets = wb.sheetnames

        try:
            if len(sheets) == 1:
                sheet = sheets[0]
            else:
                self.pause = True
                self.showMessageBox.emit(sheets)
                while self.pause:
                    self.sleep(1)
                sheet = self.my_window.ChoosedSheet
                # sheet = sheets[2]
        except Exception as ex:

            messagebox = QMessageBox(
                parent=self, text='Ошибка', detailedText=str(ex))
            messagebox.setWindowTitle('Внимание!')
            messagebox.setStyleSheet(
                '.QPushButton{background-color: #444444;color: white;}')
            messagebox.show()

        name_sheet = wb[sheet]
        n = 1
        k = 1
        self.my_window.wb = Workbook()
        ws = self.my_window.wb.active
        ws.title = "Сводный перечень имущества"

        ws.row_dimensions[1].ht = 39.6
        znach = ''
        for cell in name_sheet['A']:
            if re.match(r'\d', str(cell.value)):
                if name_sheet['C' + str(n)].value is not None:
                    ws['A' + str(k)] = znach
                    ws['A' + str(k)].font = Font(size="8", name='Arial')
                    ws['A' + str(k)].alignment = Alignment(horizontal='center',
                                                           vertical='center')
                    k = k + 1
            else:
                if name_sheet['C' + str(n)].value is None:
                    znach = cell.value
                else:
                    znach = cell.value
                    ws['A' + str(k)] = znach
                    ws['A' + str(k)].font = Font(size="8", name='Arial')
                    ws['A' + str(k)].alignment = Alignment(horizontal='center',
                                                           vertical='center')
                    k = k + 1
            n = n + 1
        self.my_window.ui.statusbar.showMessage('Конвертация данных')
        n = 1
        k = 1
        for cell in name_sheet['C']:
            if name_sheet['C' + str(n)].value is not None:
                ws['B' + str(k)] = cell.value
                ws['B' + str(k)].font = Font(size="8", name='Arial')
                ws['B' + str(k)].alignment = Alignment(horizontal='center',
                                                       vertical='center')
                k = k + 1
            n = n + 1
        n = 1
        k = 1
        for cell in name_sheet['J']:
            if name_sheet['C' + str(n)].value is not None:
                ws['D' + str(k)] = cell.value
                ws['D' + str(k)].font = Font(size="8", name='Arial')
                ws['D' + str(k)].alignment = Alignment(horizontal='center',
                                                       vertical='center')
                ws['D' + str(k)].number_format = '0'
                k = k + 1
            n = n + 1
        n = 1
        k = 1
        for cell in name_sheet['P']:
            if name_sheet['C' + str(n)].value is not None:
                ws['E' + str(k)] = cell.value
                ws['E' + str(k)].font = Font(size="8", name='Arial')
                ws['E' + str(k)].alignment = Alignment(horizontal='center',
                                                       vertical='center')
                k = k + 1
            n = n + 1
        n = 1
        k = 1
        for cell in name_sheet['V']:
            if name_sheet['C' + str(n)].value is not None:
                ws['C' + str(k)] = cell.value
                ws['C' + str(k)].font = Font(size="8", name='Arial')
                ws['C' + str(k)].alignment = Alignment(horizontal='center',
                                                       vertical='center')
                k = k + 1
            n = n + 1
        n = 1
        k = 1
        for cell in ws['B']:
            ws['F' + str(k)] = "=YEARFRAC(E" + str(k) + ",TODAY(),1)"
            ws['F' + str(k)].font = Font(size="8", name='Arial')
            ws['F' + str(k)].alignment = Alignment(horizontal='center',
                                                   vertical='center')
            k = k + 1
            n = n + 1
        self.my_window.ui.statusbar.showMessage('Последние штрихи')
        n = 1
        k = 1
        for cell in ws['B']:
            ws['H' + str(k)] = "=IF((G" + str(k) + "-F" + str(k) + ")<0,F" + str(k) + "-G" + str(k) + ",\"в пределах " \
                                                                                                      "срока\") "
            ws['H' + str(k)].font = Font(size="8", name='Arial')
            ws['H' + str(k)].alignment = Alignment(horizontal='center',
                                                   vertical='center')
            ws['H' + str(k)].number_format = '0.0'

            for key in lifetime:
                if key.lower() in str(cell.value).lower():
                    ws['G' + str(k)] = lifetime[key]

            ws['G' + str(k)].font = Font(size="8", name='Arial')
            ws['G' + str(k)].alignment = Alignment(horizontal='center',
                                                   vertical='center')
            ws['G' + str(k)].number_format = '0'

            k = k + 1
            n = n + 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 48
        ws.column_dimensions['C'].width = 7
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 14.5
        ws.column_dimensions['F'].width = 14.5
        ws.column_dimensions['H'].width = 24

        ws['A1'] = 'Отдел'
        ws['A1'].font = Font(bold=True, size="10", name='Arial')
        ws['A1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['B1'] = 'Наименование имущества'
        ws['B1'].font = Font(bold=True, size="10", name='Arial')
        ws['B1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['C1'] = 'Кол-во'
        ws['C1'].font = Font(bold=True, size="10", name='Arial')
        ws['C1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['D1'] = 'Инвентарный номер'
        ws['D1'].font = Font(bold=True, size="10", name='Arial')
        ws['D1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['E1'] = 'Дата принятия к учету'
        ws['E1'].font = Font(bold=True, size="10", name='Arial')
        ws['E1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['F1'] = 'Срок использования'
        ws['F1'].font = Font(bold=True, size="10", name='Arial')
        ws['F1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['G1'] = 'Срок по нормам, лет'
        ws['G1'].font = Font(bold=True, size="10", name='Arial')
        ws['G1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['H1'] = 'Срок превышен, лет'
        ws['H1'].font = Font(bold=True, size="10", name='Arial')
        ws['H1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')

        self.my_window.ui.tableWidget.setColumnCount(ws.max_column)
        self.my_window.ui.tableWidget.setHorizontalHeaderLabels(
            [str(ws['A1'].value), str(ws['B1'].value), str(ws['C1'].value),
             str(ws['D1'].value), str(ws['E1'].value), str(ws['F1'].value),
             str(ws['G1'].value), str(ws['H1'].value)])

        self.my_window.ui.tableWidget.setRowCount(ws.max_row)

        schet = 0
        for cell in ws['A']:
            self.my_window.ui.tableWidget.setItem(
                schet, 0, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
            if cell.value != None:
                self.mnozh.add(cell.value)
        schet = 0
        for cell in ws['B']:
            self.my_window.ui.tableWidget.setItem(
                schet, 1, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0
        for cell in ws['C']:
            self.my_window.ui.tableWidget.setItem(
                schet, 2, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0
        for cell in ws['D']:
            self.my_window.ui.tableWidget.setItem(
                schet, 3, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0
        for cell in ws['E']:
            self.my_window.ui.tableWidget.setItem(
                schet, 4, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0

        for cell in ws['F']:
            self.my_window.ui.tableWidget.setItem(
                schet, 5, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0
        for cell in ws['G']:
            self.my_window.ui.tableWidget.setItem(
                schet, 6, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0
        for cell in ws['H']:
            self.my_window.ui.tableWidget.setItem(
                schet, 7, QTableWidgetItem(str(cell.value)))
            schet = schet + 1

        self.my_window.ui.tableWidget.resizeColumnsToContents()

        self.my_window.wb.save(
            'C:\Windows\Temp\Сводный перечень имущества.xlsx')

        ws.auto_filter.ref = ws.dimensions
        self.my_window.ui.statusbar.showMessage('Таблица сконвертирована')
        self.my_window.ui.pushButton_2.setEnabled(True)
        self.my_window.ui.pushButton_3.setEnabled(True)
        self.my_window.ui.label_animation.setMovie(None)
        self.my_window.movie.stop()
        self.send_mnozh.emit(self.mnozh)
        self.fihishThread.emit("ended")


class MyWindow(QtWidgets.QMainWindow):

    def __init__(self):
        self.otdels = []
        self.analizes = analize.Analyze(my_window=self)
        super(MyWindow, self).__init__()
        self.chooseFilter = ChooseFilter(my_window=self)
        self.chooseOtdelFilter = ChooseOtdelFilter(my_window=self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.aboutForm = None
        self.setWindowIcon(QtGui.QIcon(':roskazna.png'))
        self.ui.label_animation = QLabel(self)
        self.ui.label_animation.setFixedHeight(130)
        self.ui.label_animation.setFixedWidth(70)
        self.ui.label_animation.move(int(self.width() * 0.5) - int(self.ui.label_animation.width() * 0.5),
                                     int(self.height() * 0.5) - int(self.ui.label_animation.height() * 0.5))
        self.ui.pushButton_2.clicked.connect(self.btn_clicked)
        self.ui.pushButton_3.clicked.connect(self.save_btn_clicked)
        self.ui.menu.actions()[0].triggered.connect(self.OpenAbout)
        self.ui.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.ui.tableWidget.horizontalHeader().setStretchLastSection(True)
        stylesheet = "::section{background-color:rgb(252, 246, 5);}"
        self.ui.tableWidget.horizontalHeader().setStyleSheet(stylesheet)
        self.ui.tableWidget.verticalHeader().setStyleSheet(stylesheet)
        self.ui.tableWidget.verticalScrollBar().setStyleSheet('background: #444444')
        self.ui.tableWidget.horizontalScrollBar().setStyleSheet('background: #444444')
        self.ui.tableWidget_2.verticalScrollBar().setStyleSheet('background: #444444')
        self.ui.tableWidget_2.horizontalScrollBar().setStyleSheet('background: #444444')
        self.ui.tableWidget_2.horizontalHeader().setStyleSheet(stylesheet)
        self.ui.tableWidget_2.verticalHeader().setStyleSheet(stylesheet)
        self.ui.tableWidget_2.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.ui.tableWidget_2.horizontalHeader().setStretchLastSection(True)
        self.ui.pushButton.clicked.connect(self.openChooseFilter)
        self.ui.pushButton_5.clicked.connect(self.resetAnalyz)
        self.ui.pushButton_6.clicked.connect(self.openChooseOtdelFilters)
        self.ui.pushButton_7.clicked.connect(self.startAnalyz)

    def resizeEvent(self, event):
        self.ui.label_animation.move(int(self.width() * 0.5) - int(self.ui.label_animation.width() * 0.5),
                                     int(self.height() * 0.5) - int(self.ui.label_animation.height() * 0.5))
        QtWidgets.QMainWindow.resizeEvent(self, event)

    @QtCore.pyqtSlot(list)
    def msgBox(self, list):
        cDialog = ChooseWindow(list)
        if cDialog.exec():
            self.ChoosedSheet = cDialog.ChoosedSheet
            self.my_thread.pause = False
        else:
            self.ui.statusbar.showMessage('Анализ отменен')
            self.my_thread.quit()

            self.ui.pushButton_2.setEnabled(True)
            self.ui.pushButton_3.setEnabled(True)
            self.ui.label_animation.setMovie(None)
            self.movie.stop()

    def new_thread(self):
        self.my_thread = MyThread(my_window=self)
        self.my_thread.showMessageBox.connect(self.msgBox)
        self.my_thread.fihishThread.connect((self.lblDis))
        self.my_thread.send_mnozh.connect(self.chooseOtdelFilter.getMnozh)
        self.my_thread.start()
        self.movie = QMovie(':Spinner.gif')
        self.movie.setScaledSize(QSize(70, 130))
        self.ui.label_animation.setMovie(self.movie)
        self.movie.start()

    @QtCore.pyqtSlot(str)
    def lblDis(self, str):
        self.ui.label_animation.setEnabled(False)

    def btn_clicked(self):
        self.filename = QFileDialog.getOpenFileName(
            None, 'Открыть', os.path.dirname("C:\\"), 'All Files(*.xlsx)')
        if str(self.filename) in "('', '')":
            self.ui.statusbar.showMessage('Файл не выбран')
        else:
            self.new_thread()

    def save_btn_clicked(self):
        file_save, _ = QFileDialog.getSaveFileName(
            self, 'Сохранить', 'Сводный перечень имущества', 'All Files(*.xlsx)')
        try:
            if str(file_save) != "":
                self.wb.save(file_save)
                self.ui.statusbar.showMessage('Таблица сохранена')
        except PermissionError as err:
            messagebox = QMessageBox(
                parent=self, text='Ошибка доступа. Необходимо закрыть файл', detailedText=str(err))
            messagebox.setWindowTitle('Внимание!')
            messagebox.setStyleSheet(
                '.QPushButton{background-color: #444444;color: white;}')
            messagebox.show()
        except Exception as ex:
            messagebox = QMessageBox(
                parent=self, text='Ошибка', detailedText=str(ex))
            messagebox.setWindowTitle('Внимание!')
            messagebox.setStyleSheet(
                '.QPushButton{background-color: #444444;color: white;}')
            messagebox.show()

    def OpenAbout(self):
        if (self.aboutForm != None):
            self.aboutForm.close()
        self.aboutForm = AboutWindows()
        self.aboutForm.show()

    def openChooseFilter(self):
        self.chooseFilter.show()

    def openChooseOtdelFilters(self):
        self.chooseOtdelFilter.show()

    def resetAnalyz(self):
        self.ui.tableWidget_2.clear()
        self.ui.tableWidget_2.setRowCount(0)
        self.ui.tableWidget_2.setColumnCount(0)

    def startAnalyz(self):
        self.filename = 'C:\Windows\Temp\Сводный перечень имущества.xlsx'
        self.otdels = self.analizes.analyze_xls(filename=self.filename)

        for i in range(self.ui.tableWidget_2.rowCount()):
            for j in range(self.ui.tableWidget_2.columnCount()):
                for otdel in self.otdels:
                    if(otdel.name == self.ui.tableWidget_2.verticalHeaderItem(i).text()):
                        tempFlag = False
                        for ship in otdel.shipments:
                            if(ship.name == self.ui.tableWidget_2.horizontalHeaderItem(j).text()):
                                self.ui.tableWidget_2.setItem(
                                    i, j, QTableWidgetItem(str(ship.shipCount)))
                                tempFlag = True
                            if(choose_position_header_evry_two[ship.name] == self.ui.tableWidget_2.horizontalHeaderItem(j).text()):
                                self.ui.tableWidget_2.setItem(
                                    i, j, QTableWidgetItem(str(ship.expiredShipCount)))
                                tempFlag = True

                        if(tempFlag == False):
                            self.ui.tableWidget_2.setItem(
                                i, j, QTableWidgetItem(str(0)))


class ChooseFilter(QtWidgets.QDialog):

    def __init__(self, my_window):
        super(ChooseFilter, self).__init__()
        self.vivod_header = []
        self.my_window = my_window
        self.ui = Ui_Dialog_ChooseFilter()
        self.ui.setupUi(self)
        spisok = []
        self.vivod_dict = {}
        for key, value in choose_position.items():
            spisok.append(value)
        self.ui.listWidget.addItems(spisok)
        self.ui.listWidget.setSelectionMode(
            QtWidgets.QAbstractItemView.ExtendedSelection
        )

        self.ui.pushButton.clicked.connect(self.set_header_table2)
        self.ui.listWidget.verticalScrollBar().setStyleSheet('background: #444444')
        self.ui.listWidget.horizontalScrollBar().setStyleSheet('background: #444444')

    def set_header_table2(self):
        items = self.ui.listWidget.selectedItems()
        self.znach = []
        self.vivod_header = []
        for i in range(len(items)):
            self.znach.append(
                str(self.ui.listWidget.selectedItems()[i].text()))

        znach = self.znach
        for key, val in choose_position_header.items():
            for value in znach:
                if(value == val):
                    self.vivod_header.append(key)
                    self.vivod_header.append(choose_position_header_evry_two[key])
                
        self.my_window.ui.tableWidget_2.setColumnCount(len(self.vivod_header))
        self.my_window.ui.tableWidget_2.setHorizontalHeaderLabels(
            self.vivod_header)
        self.my_window.ui.tableWidget_2.resizeColumnsToContents()
        self.my_window.analizes.set_znach(self.znach)

        self.close()


class ChooseOtdelFilter(QtWidgets.QDialog):
    def __init__(self, my_window):
        super(ChooseOtdelFilter, self).__init__()
        self.my_window = my_window
        self.ui = Ui_Dialog_ChooseFilter()
        self.ui.setupUi(self)

        self.ui.listWidget.setSelectionMode(
            QtWidgets.QAbstractItemView.ExtendedSelection
        )
        self.ui.pushButton.clicked.connect(self.set_header_table2)
        self.ui.listWidget.verticalScrollBar().setStyleSheet('background: #444444')
        self.ui.listWidget.horizontalScrollBar().setStyleSheet('background: #444444')

    @QtCore.pyqtSlot(set)
    def getMnozh(self, set):
        self.mnozh = set
        self.sorted_list = list(self.mnozh)
        self.sorted_list.sort()
        self.ui.listWidget.addItems(self.sorted_list)

    def set_header_table2(self):
        items = self.ui.listWidget.selectedItems()
        self.otdel = []
        self.my_window.otdels = []
        for i in range(len(items)):
            self.otdel.append(
                str(self.ui.listWidget.selectedItems()[i].text()))
            self.my_window.otdels.append(
                Otdel(self.ui.listWidget.selectedItems()[i].text()))

        self.my_window.ui.tableWidget_2.setRowCount(len(self.otdel))
        self.my_window.ui.tableWidget_2.setVerticalHeaderLabels(self.otdel)
        self.my_window.ui.tableWidget_2.resizeColumnsToContents()
        self.my_window.ui.tableWidget_2.verticalHeader(
        ).setSectionResizeMode(QHeaderView.Stretch)

        self.my_window.analizes.set_otdel(otdel=self.my_window.otdels)
        self.close()


class AboutWindows(QtWidgets.QDialog):

    def __init__(self):
        super(AboutWindows, self).__init__()
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)


class ChooseWindow(QtWidgets.QDialog):
    def __init__(self, list):
        super(ChooseWindow, self).__init__()
        self.list = list
        self.ui = Choose_Dialog()
        self.ui.setupUi(self)
        self.ui.buttonBox.accepted.connect(self.ChooseList)
        self.ui.buttonBox.rejected.connect(self.RejectChooseList)
        self.setWindowTitle("Выбор листа для анализа")

        for value in self.list:
            self.ui.comboBox.addItem(str(value))

    def ChooseList(self):
        self.ChoosedSheet = self.ui.comboBox.currentText()

    def RejectChooseList(self):
        self.ChoosedSheet = 'None'


app = QtWidgets.QApplication([])
application = MyWindow()
application.setWindowTitle("Конвертер ведомости учета имущества")
application.show()

sys.exit(app.exec())
