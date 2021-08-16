from PyQt5 import QtCore
from PyQt5.QtCore import QThread
from PyQt5.QtWidgets import QTableWidgetItem, QMessageBox

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from classes.dicts import lifetime
from classes.dicts import spravochnik

import re


class MyThread(QThread):
    showMessageBox = QtCore.pyqtSignal(list)
    fihishThread = QtCore.pyqtSignal(str)
    send_mnozh = QtCore.pyqtSignal(set)

    def __init__(self, my_window, parent=None):
        super(MyThread, self).__init__()
        self.my_window = my_window
        self.pause = False
        self.mnozh = set()
        self.number_row = 1
        
    def zapoln_A(self, ws, name_sheet):
        n = 1
        if self.my_window.first:
            k = ws.max_row
            self.number_row = k
        else:
            k = ws.max_row + 1
            self.number_row = k
            self.store = k

        znach = ''
        for cell in name_sheet['A']:
            if re.match(r'\d', str(cell.value)):
                if name_sheet['C' + str(n)].value is not None:
                    ws['A' + str(k)] = znach
                    ws['A' + str(k)].font = Font(size="8", name='Arial')
                    ws['A' + str(k)].alignment = Alignment(horizontal='center',
                                                           vertical='center')
                    k = k + 1
            else:
                if name_sheet['C' + str(n)].value is None:
                    znach = cell.value
                    if cell.value is not None:
                        for key, value in spravochnik.items():
                            if znach in key:
                                znach = value
                else:
                    znach = cell.value
                    if cell.value is not None:
                        for key, value in spravochnik.items():
                            if znach in key:
                                znach = value
                    ws['A' + str(k)] = znach
                    ws['A' + str(k)].font = Font(size="8", name='Arial')
                    ws['A' + str(k)].alignment = Alignment(horizontal='center',
                                                           vertical='center')
                    k = k + 1
            n = n + 1

    def zapoln_B(self, ws, name_sheet):
        n = 1
        k = self.number_row
  
        for cell in name_sheet['C']:
            if name_sheet['C' + str(n)].value is not None:
                ws['B' + str(k)] = cell.value
                ws['B' + str(k)].font = Font(size="8", name='Arial')
                ws['B' + str(k)].alignment = Alignment(horizontal='center',
                                                       vertical='center')
                k = k + 1
            n = n + 1

    def zapoln_D(self, ws, name_sheet):
        n = 1
        k = self.number_row    
        
        for cell in name_sheet['J']:
            if name_sheet['C' + str(n)].value is not None:
                ws['D' + str(k)] = cell.value
                ws['D' + str(k)].font = Font(size="8", name='Arial')
                ws['D' + str(k)].alignment = Alignment(horizontal='center',
                                                       vertical='center')
                ws['D' + str(k)].number_format = '0'
                k = k + 1
            n = n + 1

    def zapoln_E(self, ws, name_sheet):
        n = 1
        k = self.number_row 

        for cell in name_sheet['P']:
            if name_sheet['C' + str(n)].value is not None:
                ws['E' + str(k)] = cell.value
                ws['E' + str(k)].font = Font(size="8", name='Arial')
                ws['E' + str(k)].alignment = Alignment(horizontal='center',
                                                       vertical='center')
                k = k + 1
            n = n + 1

    def zapoln_C(self, ws, name_sheet):
        n = 1
        k = self.number_row 

        for cell in name_sheet['V']:
            if name_sheet['C' + str(n)].value is not None:
                ws['C' + str(k)] = cell.value
                ws['C' + str(k)].font = Font(size="8", name='Arial')
                ws['C' + str(k)].alignment = Alignment(horizontal='center',
                                                       vertical='center')
                k = k + 1
            n = n + 1

        if self.my_window.first is not True:
            ws.delete_rows(self.store, 1)

    def zapoln_F(self, ws, name_sheet):
        n = 1
        k = self.number_row 
        
        for cell in name_sheet['V']: 
            if name_sheet['C' + str(n)].value is not None:
                ws['F' + str(k)] = "=YEARFRAC(E" + str(k) + ",TODAY(),1)"
                ws['F' + str(k)].font = Font(size="8", name='Arial')
                ws['F' + str(k)].alignment = Alignment(horizontal='center',
                                                    vertical='center')
                k = k + 1
            n = n + 1

    def zapoln_H_I_J_K_G(self, ws, name_sheet):
        n = 1
        k = self.number_row

        for cell in name_sheet['V']:
            if name_sheet['C' + str(n)].value is not None:
                ws['H' + str(k)] = "=IF((G" + str(k) + "-F" + str(k) + ")<0,F" + str(k) + "-G" + str(k) + ",\"в пределах " \
                                                                                                        "срока\") "
                ws['H' + str(k)].font = Font(size="8", name='Arial')
                ws['H' + str(k)].alignment = Alignment(horizontal='center',
                                                    vertical='center')
                ws['H' + str(k)].number_format = '0.0'

                ws['I' + str(k)] = "=E" + str(k) + "+" + "G" + str(k) + "*365"
                ws['I' + str(k)].font = Font(size="8", name='Arial')
                ws['I' + str(k)].alignment = Alignment(horizontal='center',
                                                        vertical='center')
                ws['I' + str(k)].number_format = 'DD.mm.YYYY'

                ws['J' + str(k)] = "=TEXT(I" + str(k) + ",\"гггг\")"
                ws['J' + str(k)].font = Font(size="8", name='Arial')
                ws['J' + str(k)].alignment = Alignment(horizontal='center',
                                                        vertical='center')

                ws['K' + str(k)] = "=IF(J" + str(k) + "<=\"2022\""  + ",\"да\"" + ",\"нет\")"
                ws['K' + str(k)].font = Font(size="8", name='Arial')
                ws['K' + str(k)].alignment = Alignment(horizontal='center',
                                                        vertical='center')
                

                for key in lifetime:
                    if key.lower() in str(ws['B' + str(k)].value).lower():
                        ws['G' + str(k)] = lifetime[key]

                ws['G' + str(k)].font = Font(size="8", name='Arial')
                ws['G' + str(k)].alignment = Alignment(horizontal='center',
                                                    vertical='center')
                ws['G' + str(k)].number_format = '0'

                k = k + 1
            n = n + 1

        if self.my_window.first is not True:
            ws.delete_rows(ws.max_row, 1)

    def style_cell(self, ws):
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 48
        ws.column_dimensions['C'].width = 7
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 14.5
        ws.column_dimensions['F'].width = 14.5
        ws.column_dimensions['H'].width = 24
        ws.column_dimensions['I'].width = 24
        ws.column_dimensions['J'].width = 14.5
        ws.column_dimensions['K'].width = 14.5
        ws.row_dimensions[1].ht = 39.6

        ws['A1'] = 'Отдел'
        ws['A1'].font = Font(bold=True, size="10", name='Arial')
        ws['A1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['B1'] = 'Наименование имущества'
        ws['B1'].font = Font(bold=True, size="10", name='Arial')
        ws['B1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['C1'] = 'Кол-во'
        ws['C1'].font = Font(bold=True, size="10", name='Arial')
        ws['C1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['D1'] = 'Инвентарный номер'
        ws['D1'].font = Font(bold=True, size="10", name='Arial')
        ws['D1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['E1'] = 'Дата ввода в эксплуатацию'
        ws['E1'].font = Font(bold=True, size="10", name='Arial')
        ws['E1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['F1'] = 'Срок использования'
        ws['F1'].font = Font(bold=True, size="10", name='Arial')
        ws['F1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['G1'] = 'Срок по нормам, лет'
        ws['G1'].font = Font(bold=True, size="10", name='Arial')
        ws['G1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['H1'] = 'Срок превышен, лет'
        ws['H1'].font = Font(bold=True, size="10", name='Arial')
        ws['H1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['I1'] = 'Срок годности для неистекших (из расчета 365 дней в году)'
        ws['I1'].font = Font(bold=True, size="10", name='Arial')
        ws['I1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['J1'] = 'Год истечения срока годности'
        ws['J1'].font = Font(bold=True, size="10", name='Arial')
        ws['J1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        ws['K1'] = 'Истекает к концу 2022 года'
        ws['K1'].font = Font(bold=True, size="10", name='Arial')
        ws['K1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')

        tumbler = 0
        for row_cells in ws.iter_rows():
            tumbler = 0
            for cell in row_cells:
                if cell.value == 'Аппарат Управления':
                    tumbler = 1
                if cell.value == 'Склад':
                    tumbler = 2
                if tumbler == 1:
                    my_red = openpyxl.styles.colors.Color(rgb='00B7DEE8')
                    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
                    cell.fill = my_fill   
                if tumbler == 2:
                    my_red = openpyxl.styles.colors.Color(rgb='00EBF1DE')
                    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
                    cell.fill = my_fill  

    def table_widget_zapoln(self, ws):

        self.my_window.ui.tableWidget.setColumnCount(ws.max_column)
        self.my_window.ui.tableWidget.setHorizontalHeaderLabels(
            [str(ws['A1'].value), str(ws['B1'].value), str(ws['C1'].value),
             str(ws['D1'].value), str(ws['E1'].value), str(ws['F1'].value),
             str(ws['G1'].value), str(ws['H1'].value), str(ws['I1'].value), str(ws['J1'].value), str(ws['K1'].value)])

        self.my_window.ui.tableWidget.setRowCount(ws.max_row)

        schet = 0
        for cell in ws['A']:
            self.my_window.ui.tableWidget.setItem(
                schet, 0, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
            if cell.value != None and cell.value != 'Отдел':
                self.mnozh.add(cell.value)

        schet = 0
        for cell in ws['B']:
            self.my_window.ui.tableWidget.setItem(
                schet, 1, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0
        for cell in ws['C']:
            self.my_window.ui.tableWidget.setItem(
                schet, 2, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0
        for cell in ws['D']:
            self.my_window.ui.tableWidget.setItem(
                schet, 3, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0
        for cell in ws['E']:
            self.my_window.ui.tableWidget.setItem(
                schet, 4, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0

        for cell in ws['F']:
            self.my_window.ui.tableWidget.setItem(
                schet, 5, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0
        for cell in ws['G']:
            self.my_window.ui.tableWidget.setItem(
                schet, 6, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0
        for cell in ws['H']:
            self.my_window.ui.tableWidget.setItem(
                schet, 7, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0   
        for cell in ws['I']:
            self.my_window.ui.tableWidget.setItem(
                schet, 8, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0   
        for cell in ws['J']:
            self.my_window.ui.tableWidget.setItem(
                schet, 9, QTableWidgetItem(str(cell.value)))
            schet = schet + 1
        schet = 0   
        for cell in ws['K']:
            self.my_window.ui.tableWidget.setItem(
                schet, 10, QTableWidgetItem(str(cell.value)))
            schet = schet + 1

        self.my_window.ui.tableWidget.resizeColumnsToContents()

    def run(self):
        self.my_window.ui.statusbar.showMessage(
            'Анализ и сопоставление данных исходной таблицы')
        self.my_window.ui.pushButton_2.setEnabled(False)
        self.my_window.ui.pushButton_3.setEnabled(False)
        wb = openpyxl.load_workbook(self.my_window.filename[0])
        sheets = wb.sheetnames

        try:
            if len(sheets) == 1:
                sheet = sheets[0]
            else:
                self.pause = True
                self.showMessageBox.emit(sheets)
                while self.pause:
                    self.sleep(1)
                sheet = self.my_window.ChoosedSheet
        except Exception as ex:

            messagebox = QMessageBox(
                parent=self, text='Ошибка', detailedText=str(ex))
            messagebox.setWindowTitle('Внимание!')
            messagebox.setStyleSheet(
                '.QPushButton{background-color: #444444;color: white;}')
            messagebox.show()

        name_sheet = wb[sheet]

        if self.my_window.first:

            self.my_window.wb = Workbook()
            ws = self.my_window.wb.active
            ws.title = "Сводный перечень имущества"
        
        else:
            self.my_window.wb = openpyxl.load_workbook('C:\Windows\Temp\Сводный перечень имущества.xlsx')
            ws = self.my_window.wb.active

        self.zapoln_A(ws, name_sheet)
        self.my_window.ui.statusbar.showMessage('Конвертация данных')
        self.zapoln_B(ws, name_sheet)
        self.zapoln_D(ws, name_sheet)
        self.zapoln_E(ws, name_sheet)
        self.zapoln_C(ws, name_sheet)
        self.zapoln_F(ws, name_sheet)
        self.my_window.ui.statusbar.showMessage('Последние штрихи')
        self.zapoln_H_I_J_K_G(ws, name_sheet)
        self.style_cell(ws)
        self.table_widget_zapoln(ws)

        self.my_window.wb.save(
            'C:\Windows\Temp\Сводный перечень имущества.xlsx')

        self.my_window.ui.pushButton.setEnabled(True)
        self.my_window.ui.pushButton_5.setEnabled(True)
        self.my_window.ui.pushButton_6.setEnabled(True)
        self.my_window.ui.pushButton_7.setEnabled(True)

        ws.auto_filter.ref = ws.dimensions
        self.my_window.ui.statusbar.showMessage('Таблица сконвертирована')
        self.my_window.ui.pushButton_2.setEnabled(True)
        self.my_window.ui.pushButton_3.setEnabled(True)
        self.my_window.ui.label_animation.setMovie(None)
        self.my_window.movie.stop()
        self.send_mnozh.emit(self.mnozh)
        self.fihishThread.emit("ended")
        self.my_window.first = False

