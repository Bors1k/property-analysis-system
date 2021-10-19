from ntpath import join
from PyQt5 import QtWidgets, QtGui
from PyQt5 import QtCore
from PyQt5.QtGui import QMovie
from PyQt5.QtWidgets import QFileDialog, QLabel, QTableWidgetItem, QAbstractItemView, QMessageBox

from PyQt5.QtCore import QSize

from UIforms.MainForm.form import Ui_MainWindow
from classes.forms.ChooseOtdelFilter import ChooseOtdelFilter
from classes.forms.AboutWindow import AboutWindows
from classes.forms.ChooseWindow import ChooseWindow
from classes.forms.ChooseFilter import ChooseFilter
from classes.forms.SpravWindow import SpravWindow

from classes import analize
from classes.myThread import MyThread
from classes.dicts import Dictionary


from openpyxl.styles.borders import BORDER_MEDIUM, Border, Side, BORDER_THIN
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import pymorphy2
import openpyxl
import datetime
import os
import json


class MyWindow(QtWidgets.QMainWindow):

    def __init__(self):
        self.first = True
        self.otdels = []
        self.analizes = analize.Analyze(my_window=self)
        super(MyWindow, self).__init__()
        self.dictionary = Dictionary()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.aboutForm = None
        self.spravForm = None
        self.setWindowIcon(QtGui.QIcon(':roskazna.png'))
        self.ui.label_animation = QLabel(self)
        self.ui.label_animation.setFixedHeight(130)
        self.ui.label_animation.setFixedWidth(70)
        self.ui.label_animation.move(int(self.width() * 0.5) - int(self.ui.label_animation.width() * 0.5),
                                     int(self.height() * 0.5) - int(self.ui.label_animation.height() * 0.5))
        self.ui.pushButton_2.clicked.connect(self.btn_clicked)
        self.ui.pushButton_3.clicked.connect(self.save_btn_clicked)
        self.ui.menu.actions()[0].triggered.connect(self.OpenAbout)
        self.ui.menu.actions()[1].triggered.connect(self.OpenSprav)
        self.ui.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.ui.tableWidget.horizontalHeader().setStretchLastSection(True)
        stylesheet = "::section{background-color:rgb(252, 246, 5);}"
        self.ui.tableWidget.horizontalHeader().setStyleSheet(stylesheet)
        self.ui.tableWidget.verticalHeader().setStyleSheet(stylesheet)
        self.ui.tableWidget.verticalScrollBar().setStyleSheet('background: #444444')
        self.ui.tableWidget.horizontalScrollBar().setStyleSheet('background: #444444')
        self.ui.tableWidget_2.verticalScrollBar().setStyleSheet('background: #444444')
        self.ui.tableWidget_2.horizontalScrollBar().setStyleSheet('background: #444444')
        self.ui.tableWidget_2.horizontalHeader().setStyleSheet(stylesheet)
        self.ui.tableWidget_2.verticalHeader().setStyleSheet(stylesheet)
        self.ui.tableWidget_2.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.ui.tableWidget_2.horizontalHeader().setStretchLastSection(True)
        self.ui.pushButton.clicked.connect(self.openChooseFilter)
        self.ui.pushButton_5.clicked.connect(self.resetAnalyz)
        self.ui.pushButton_6.clicked.connect(self.openChooseOtdelFilters)
        self.ui.pushButton_7.clicked.connect(self.startAnalyz)
        self.ui.pushButton.setEnabled(False)
        self.ui.pushButton_5.setEnabled(False)
        self.ui.pushButton_6.setEnabled(False)
        self.ui.pushButton_7.setEnabled(False)
        self.dictionary.zapoln_dict()

    def resizeEvent(self, event):
        self.ui.label_animation.move(int(self.width() * 0.5) - int(self.ui.label_animation.width() * 0.5),
                                     int(self.height() * 0.5) - int(self.ui.label_animation.height() * 0.5))
        QtWidgets.QMainWindow.resizeEvent(self, event)

    @QtCore.pyqtSlot(list)
    def msgBox(self, list):
        cDialog = ChooseWindow(list)
        if cDialog.exec():
            self.ChoosedSheet = cDialog.ChoosedSheet
            self.my_thread.pause = False
        else:
            self.ui.statusbar.showMessage('Анализ отменен')
            self.my_thread.quit()

            self.ui.pushButton_2.setEnabled(True)
            self.ui.pushButton_3.setEnabled(True)
            self.ui.label_animation.setMovie(None)
            self.movie.stop()

    def new_thread(self):
        self.chooseOtdelFilter = ChooseOtdelFilter(my_window=self)
        self.my_thread = MyThread(my_window=self)
        self.my_thread.showMessageBox.connect(self.msgBox)
        self.my_thread.fihishThread.connect((self.lblDis))
        self.my_thread.send_mnozh.connect(self.chooseOtdelFilter.getMnozh)
        self.my_thread.start()
        self.movie = QMovie(':Spinner.gif')
        self.movie.setScaledSize(QSize(70, 130))
        self.ui.label_animation.setMovie(self.movie)
        self.movie.start()

    @QtCore.pyqtSlot(str)
    def lblDis(self, str):
        self.ui.label_animation.setEnabled(False)

    def btn_clicked(self):
        with open ("C:\\Users\\Public\\property-analysis-system\\dicts.json", encoding='utf-8') as f:
            templates = json.load(f)
            open_folder = list(templates["open_folder"])
            for i in range(len(open_folder)):
                if open_folder[-1] != '\\' and open_folder[-1] != 'C:\\':
                    del open_folder[-1]
            open_folder = ''.join(open_folder)
        self.filename = QFileDialog.getOpenFileName(
            None, 'Открыть', open_folder, 'All Files(*.xlsx)')
        if str(self.filename) in "('', '')":
            self.ui.statusbar.showMessage('Файл не выбран')
        else:
            filename = self.filename[0]
            filename = filename.replace('/', '\\')
            filename = filename.replace('.xlsx', '')
            with open('C:\\Users\\Public\\property-analysis-system\\dicts.json', 'r+') as f:
                json_data = json.load(f)
                json_data['open_folder'] = filename
                f.seek(0)
                f.write(json.dumps(json_data))
                f.truncate()

            self.new_thread()

    def save_btn_clicked(self):
        with open ("C:\\Users\\Public\\property-analysis-system\\dicts.json", encoding='utf-8') as f:
            templates = json.load(f)
            save_folder = list(templates["save_folder"])
            save_folder = ''.join(save_folder)
            file_save, _ = QFileDialog.getSaveFileName(
                self, 'Сохранить', save_folder, 'Excel(*.xlsx)')
        try:
            if str(file_save) != "":
                self.wb.save(file_save)
                self.ui.statusbar.showMessage('Таблица сохранена')
                file_save = file_save.replace('/', '\\')
                file_save = file_save.replace('.xlsx', '')
                with open('C:\\Users\\Public\\property-analysis-system\\dicts.json', 'r+') as f:
                    json_data = json.load(f)
                    json_data['save_folder'] = file_save
                    f.seek(0)
                    f.write(json.dumps(json_data))
                    f.truncate()

        except PermissionError as err:
            messagebox = QMessageBox(
                parent=self, text='Ошибка доступа. Необходимо закрыть файл', detailedText=str(err))
            messagebox.setWindowTitle('Внимание!')
            messagebox.setStyleSheet(
                '.QPushButton{background-color: #444444;color: white;}')
            messagebox.show()
        except Exception as ex:
            messagebox = QMessageBox(
                parent=self, text='Ошибка', detailedText=str(ex))
            messagebox.setWindowTitle('Внимание!')
            messagebox.setStyleSheet(
                '.QPushButton{background-color: #444444;color: white;}')
            messagebox.show()

    def OpenAbout(self):
        if (self.aboutForm != None):
            self.aboutForm.close()
        self.aboutForm = AboutWindows()
        self.aboutForm.show()

    def add_dict_json(self):
        imushestvo = self.spravForm.ui.textEdit.toPlainText()
        srok_po_normam = int(self.spravForm.ui.textEdit_2.toPlainText())
        print(imushestvo)
        print(srok_po_normam)
        morph = pymorphy2.MorphAnalyzer()
        res = morph.parse(imushestvo)[0]
        res.inflect({'plur', 'gent'}) 
        print(res.inflect({'plur', 'gent'}).word)
        new_per = res.inflect({'gent'}).word
        print(new_per)

        with open ("C:\\Users\\Public\\property-analysis-system\\dicts.json", encoding='utf-8') as f:
            templates = json.load(f)
            lifetime = dict(templates['lifetime'])
            choose_position = dict(templates['choose_position'])
            choose_position_header = dict(templates['choose_position_header'])
            choose_position_header_evry_two = dict(templates['choose_position_header_evry_two'])
            
        choose_position_header_evry_two['Количество ' + res.inflect({'plur', 'gent'}).word] = 'Из них ' + res.inflect({'plur'}).word + ' с превышенным сроком'
        choose_position_header['Количество ' + res.inflect({'plur', 'gent'}).word] = imushestvo
        choose_position[imushestvo] = imushestvo
        lifetime[imushestvo] = srok_po_normam
        with open('C:\\Users\\Public\\property-analysis-system\\dicts.json', 'r+') as f:
            json_data = json.load(f)
            json_data['lifetime'] = lifetime
            json_data['choose_position'] = choose_position
            json_data['choose_position_header'] = choose_position_header
            json_data['choose_position_header_evry_two'] = choose_position_header_evry_two
            f.seek(0)
            f.write(json.dumps(json_data))
            f.truncate()
        self.OpenSprav()
        # dicts.zapoln_dict()

    def del_dict_json(self):
        row = self.spravForm.ui.tableNorm.currentItem().row()
        cell =self.spravForm.ui.tableNorm.item(row, 0).text() 
        morph = pymorphy2.MorphAnalyzer()
        res = morph.parse(cell)[0]
        with open ("C:\\Users\\Public\\property-analysis-system\\dicts.json", encoding='utf-8') as f:
            templates = json.load(f)
            lifetime = dict(templates['lifetime'])            
            del lifetime[cell]

            choose_position = dict(templates['choose_position'])
            del choose_position[cell]

            choose_position_header = dict(templates['choose_position_header'])
            del choose_position_header['Количество ' + res.inflect({'plur', 'gent'}).word]

            choose_position_header_evry_two = dict(templates['choose_position_header_evry_two'])
            del choose_position_header_evry_two['Количество ' + res.inflect({'plur', 'gent'}).word]

        with open('C:\\Users\\Public\\property-analysis-system\\dicts.json', 'r+') as f:
            json_data = json.load(f)
            json_data['lifetime'] = lifetime
            json_data['choose_position'] = choose_position
            json_data['choose_position_header'] = choose_position_header
            json_data['choose_position_header_evry_two'] = choose_position_header_evry_two
            f.seek(0)
            f.write(json.dumps(json_data))
            f.truncate()
        self.OpenSprav()
        
    def OpenSprav(self):
        if (self.spravForm != None):
            self.spravForm.close()
        self.spravForm = SpravWindow()
        self.spravForm.ui.pushButton.clicked.connect(self.add_dict_json)
        self.spravForm.ui.pushButton_2.clicked.connect(self.del_dict_json)
        stylesheet = "::section{background-color: #444444;}"
        self.spravForm.ui.tableNorm.horizontalHeader().setStyleSheet(stylesheet)
        self.spravForm.ui.tableNorm.verticalHeader().setStyleSheet(stylesheet)
        self.spravForm.ui.tableNorm.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.spravForm.ui.tableNorm.verticalScrollBar().setStyleSheet('background: #444444')
        self.spravForm.ui.tableNorm.horizontalScrollBar().setStyleSheet('background: #444444')
        self.spravForm.ui.centralwidget.setStyleSheet('.QTableWidget .QTableCornerButton::section {background-color: rgba(0,0,0,0);')
        with open ("C:\\Users\\Public\\property-analysis-system\\dicts.json", encoding='utf-8') as f:
            templates = json.load(f)
            lifetime = dict(templates['lifetime'])
            # print(lifetime)
            
            self.spravForm.ui.tableNorm.setColumnCount(2)
            self.spravForm.ui.tableNorm.setHorizontalHeaderLabels(
            ['Имущество', 'Срок по нормам'])

            self.spravForm.ui.tableNorm.setRowCount(len(lifetime))

            schet = 0
            for key, value in lifetime.items():
                # print(value)
                self.spravForm.ui.tableNorm.setItem(
                    schet, 0, QTableWidgetItem(key))
                self.spravForm.ui.tableNorm.setItem(
                    schet, 1, QTableWidgetItem(str(value)))
                schet = schet + 1

        self.spravForm.show()

    def openChooseFilter(self):
        chooseFilter = ChooseFilter(my_window=self)
        chooseFilter.exec_()

    def openChooseOtdelFilters(self):
        # chooseOtdelFilter = ChooseOtdelFilter(my_window=self)
        self.chooseOtdelFilter.exec_()

    def resetAnalyz(self):
        self.ui.tableWidget_2.clear()
        self.ui.tableWidget_2.setRowCount(0)
        self.ui.tableWidget_2.setColumnCount(0)

    def startAnalyz(self):
        self.dictionary.zapoln_dict()
        self.filename = 'C:\Windows\Temp\Сводный перечень имущества.xlsx'
        self.otdels = self.analizes.analyze_xls(filename=self.filename)

        for i in range(self.ui.tableWidget_2.rowCount()):
            for j in range(self.ui.tableWidget_2.columnCount()):
                for otdel in self.otdels:
                    if(otdel.name == self.ui.tableWidget_2.verticalHeaderItem(i).text()):
                        tempFlag = False
                        for ship in otdel.shipments:
                            if(ship.name == self.ui.tableWidget_2.horizontalHeaderItem(j).text()):
                                self.ui.tableWidget_2.setItem(
                                    i, j, QTableWidgetItem(str(ship.shipCount)))
                                tempFlag = True
                            if(self.dictionary.choose_position_header_evry_two[ship.name] == self.ui.tableWidget_2.horizontalHeaderItem(j).text()):
                                self.ui.tableWidget_2.setItem(
                                    i, j, QTableWidgetItem(str(ship.expiredShipCount)))
                                tempFlag = True
                            if((self.dictionary.choose_position_header_evry_two[ship.name] + " в " + str(datetime.date.today().year + 1)  + " году") == self.ui.tableWidget_2.horizontalHeaderItem(j).text()):
                                self.ui.tableWidget_2.setItem(
                                    i, j, QTableWidgetItem(str(ship.expiredInNextYearCount)))
                                tempFlag = True
                        if(tempFlag == False):
                            self.ui.tableWidget_2.setItem(
                                i, j, QTableWidgetItem(str(0)))

        if(len(self.wb.sheetnames) == 1):
            sheet = self.wb.create_sheet('Аналитика по отделам')
        else:
            sheet = self.wb[self.wb.sheetnames[1]]

        k = 0.9
        maxWidth = 0

        all_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_MEDIUM, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        all_border_down_r = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_MEDIUM, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_MEDIUM, color='00000000')
        )

        all_border_small = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        all_border_left = Border(
            left=Side(border_style=BORDER_MEDIUM, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        all_border_left_down = Border(
            left=Side(border_style=BORDER_MEDIUM, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_MEDIUM, color='00000000')
        )

        right_border = Border(
            left=Side(border_style=None, color='00000000'),
            right=Side(border_style=BORDER_MEDIUM, color='00000000'),
            top=Side(border_style=None, color='00000000'),
            bottom=Side(border_style=None, color='00000000')
        )

        right_bottom_border = Border(
            left=Side(border_style=None, color='00000000'),
            right=Side(border_style=BORDER_MEDIUM, color='00000000'),
            top=Side(border_style=None, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        right_bottom_border_medium = Border(
            left=Side(border_style=None, color='00000000'),
            right=Side(border_style=BORDER_MEDIUM, color='00000000'),
            top=Side(border_style=None, color='00000000'),
            bottom=Side(border_style=BORDER_MEDIUM, color='00000000')
        )

        right_bottom_border_medium_down = Border(
            left=Side(border_style=None, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=None, color='00000000'),
            bottom=Side(border_style=BORDER_MEDIUM, color='00000000')
        )

        bottom_border_small = Border(
            left=Side(border_style=None, color='00000000'),
            right=Side(border_style=None, color='00000000'),
            top=Side(border_style=None, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        bottom_border_large = Border(
            left=Side(border_style=None, color='00000000'),
            right=Side(border_style=None, color='00000000'),
            top=Side(border_style=None, color='00000000'),
            bottom=Side(border_style=BORDER_MEDIUM, color='00000000')
        )

        for i in range(self.ui.tableWidget_2.rowCount()):
            cellref = sheet.cell(i+2, 1)
            text = self.ui.tableWidget_2.verticalHeaderItem(i).text()
            cellref.value = text
            cellref.font = Font(size="8", name='Arial')
            cellref.alignment = Alignment(
                horizontal='center', vertical='center')
            if(maxWidth < len(text) * k):
                maxWidth = len(text) * k
            sheet.column_dimensions[cellref.column_letter].width = maxWidth
            cellref.border = all_border

        # Колонки снизу
        z_row = sheet.max_row
        s_row = sheet.max_row + 1
        e_row = s_row + 1
        sheet['A'+ str(s_row)] = 'Суммы для справки'
        sheet['A'+ str(s_row)].border = right_bottom_border
        sheet['A'+ str(s_row)].font = Font(bold=True, size="8", name='Arial')
        sheet['A'+ str(s_row)].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        sheet.merge_cells(start_row=s_row, start_column=1, end_row=e_row, end_column=1)
        
        sheet['A1'] = 'Наименование структурного подразделения'
        sheet['A1'].border = right_bottom_border
        sheet['A1'].font = Font(bold=True, size="8", name='Arial')
        sheet['A1'].alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        for j in range(self.ui.tableWidget_2.columnCount()):
            cellref = sheet.cell(1, j+2)
            text = self.ui.tableWidget_2.horizontalHeaderItem(j).text()
            cellref.value = text
            cellref.border = bottom_border_small
            cellref.font = Font(bold=True, size="8", name='Arial')
            cellref.alignment = Alignment(
                wrap_text=True, horizontal='center', vertical='center')
            sheet.column_dimensions[cellref.column_letter].width = 8.43

        tumbler = 0
        for row_cells in sheet.iter_rows():
            tumbler = 0
            for cell in row_cells:
                if cell.value == 'Аппарат Управления':
                    tumbler = 1
                if cell.value == 'Склад':
                    tumbler = 2
                if tumbler == 1:
                    my_red = openpyxl.styles.colors.Color(rgb='00B7DEE8')
                    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
                    cell.fill = my_fill   
                if tumbler == 2:
                    my_red = openpyxl.styles.colors.Color(rgb='00EBF1DE')
                    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
                    cell.fill = my_fill  

        for i in range(self.ui.tableWidget_2.rowCount()):
            for j in range(self.ui.tableWidget_2.columnCount()):
                text = str(self.ui.tableWidget_2.item(i, j).text())
                if(text != '0'):
                    text = text[0:len(text)-2]
                cellref = sheet.cell(i+2, j+2)
                cellref.value = int(text)
                cellref.font = Font(size="8", name='Arial')
                cellref.alignment = Alignment(
                    horizontal='center', vertical='center')
                cellref.number_format = '0'

        tumbler_two = 1
        finish_formula = "=AVERAGE("
        for j in range(self.ui.tableWidget_2.columnCount()):    
            text = self.ui.tableWidget_2.horizontalHeaderItem(j).text()
            if 'Количество' in text:
                edited_text = text.replace("Количество", "Всего")

            sheet[get_column_letter(j+2)+ str(s_row)] = "=SUM(" + get_column_letter(j+2) + "1" + ":" + get_column_letter(j+2) + str(s_row - 1)  +")"
            sheet[get_column_letter(j+2)+ str(s_row)].font = Font(size="8", name='Arial')
            sheet[get_column_letter(j+2)+ str(s_row)].alignment = Alignment(horizontal='center', vertical='center')
            sheet[get_column_letter(j+2)+ str(s_row)].font = Font(bold=True, size="8", name='Arial')
            sheet[get_column_letter(1)+ str(e_row + 2)] = 'Среднее значение превышения сроков эксплуатации движимого имущества %'
            if tumbler_two == 1:
                sheet[get_column_letter(j+2)+ str(e_row)] = edited_text
                summa = 0
                for s in range(z_row + 1):
                    if s > 1:
                        summa = summa + sheet[get_column_letter(j+2)+ str(s)].value
                tumbler_two = 2
            elif tumbler_two == 2:
                sheet[get_column_letter(j+2)+ str(e_row)] = 'с превыш. сроком'
                if summa == 0:
                    sheet[get_column_letter(j+2)+ str(e_row + 1)] = 0
                else:
                    sheet[get_column_letter(j+2)+ str(e_row + 1)] = "=SUM(" + get_column_letter(j+2) + "1" + ":" + get_column_letter(j+2) + str(s_row - 1)  +")/" + get_column_letter(j+1) + str(s_row) + "* 100"
                sheet[get_column_letter(j+2)+ str(e_row + 2)] = '% с превыш. сроком'
                tumbler_two = 3
            elif tumbler_two == 3:
                sheet[get_column_letter(j+2)+ str(e_row)] = ' срок будет превышен в 2022'
                if summa == 0:
                    sheet[get_column_letter(j+2)+ str(e_row + 1)] = 0
                else:
                    sheet[get_column_letter(j+2)+ str(e_row + 1)] = "=SUM(" + get_column_letter(j+2) + "1" + ":" + get_column_letter(j+2) + str(s_row - 1)  +")/" + get_column_letter(j) + str(s_row) + "* 100"
                sheet[get_column_letter(j+2)+ str(e_row + 2)] = '% срок будет превышен в 2022'
                finish_formula = finish_formula + get_column_letter(j+2)+ str(e_row + 1) + ','
                tumbler_two = 1

            sheet[get_column_letter(j+2)+ str(e_row)].font = Font(size="8", name='Arial')
            sheet[get_column_letter(j+2)+ str(e_row)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            sheet[get_column_letter(j+2)+ str(e_row + 2)].font = Font(size="8", name='Arial')
            sheet[get_column_letter(j+2)+ str(e_row + 2)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            sheet[get_column_letter(1)+ str(e_row + 2)].font = Font(bold=True, size="8", name='Arial')
            sheet[get_column_letter(1)+ str(e_row + 2)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            sheet[get_column_letter(j+2)+ str(e_row + 1)].font = Font(bold=True, size="8", name='Arial')
            sheet[get_column_letter(j+2)+ str(e_row + 1)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            # sheet[get_column_letter(1)+ str(e_row + 3)] = finish_formula
            sheet[get_column_letter(1)+ str(e_row + 3)].font = Font(bold=True, size="8", name='Arial')
            sheet[get_column_letter(1)+ str(e_row + 3)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        for l in range(sheet.max_column):
            if l > 2:
                sheet.cell(1, l).border = all_border_small
                sheet.cell(s_row, l).border = all_border_small
                sheet.cell(e_row, l).border = right_bottom_border_medium_down

        for s in range(s_row):

            t = 1
            for l in range(sheet.max_column + 1):
                if s > 1 and l > 1:
                    if t == 1:
                        sheet.cell(s, l).border = all_border_left
                        t = 2
                    elif t == 2:
                        sheet.cell(s, l).border = all_border_small
                        sheet.cell(e_row + 1, l).border = all_border_left
                        sheet.cell(e_row + 2, l).border = all_border_left_down
                        t = 3
                    elif t == 3:
                        sheet.cell(s, l).border = all_border
                        sheet.cell(e_row + 1, l).border = all_border
                        sheet.cell(e_row + 2, l).border = all_border_down_r
                        sheet.cell(s_row, l).border = all_border
                        sheet.cell(e_row, l).border = all_border_down_r
                    
                        t = 1

        # for l in range(sheet.max_column):
        #     if l > 2:
        #         sheet.cell(1, l).border = all_border_small
        #         sheet.cell(s_row, l).border = all_border_small
        #         sheet.cell(e_row, l).border = right_bottom_border_medium_down

        sheet.cell(s_row, sheet.max_column).border = all_border
        sheet.cell(e_row, sheet.max_column).border = right_bottom_border_medium
        sheet.cell(e_row, 2).border = all_border_left_down
        

        finish_formula = finish_formula[0:-1]
        finish_formula = finish_formula + ")"
        sheet[get_column_letter(1)+ str(e_row + 3)] = finish_formula
        sheet.row_dimensions[e_row].hight = 49
        sheet.row_dimensions[e_row + 2].hight = 95
        sheet.auto_filter.ref = sheet.dimensions

